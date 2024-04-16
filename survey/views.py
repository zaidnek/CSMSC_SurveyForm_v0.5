import re
import csv
import openpyxl
from django.shortcuts import render, redirect
from .forms import QuestionnaireForm
from .models import UserResponse
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden, StreamingHttpResponse, HttpResponse
from django.views.generic import ListView
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import IntegrityError
from django.contrib import messages


def questionnaire(request):
    if request.method == 'POST':
        # Validation pattern to only accept letters
        pattern = re.compile("^[A-Za-z ]+$")
        
        # Retrieve data from POST request
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        staff_name = request.POST.get('staff_name')
        
        # Validate first name, last name, and staff name
        if not (pattern.match(first_name) and pattern.match(last_name) and pattern.match(staff_name)):
            # Return to the form with an error message if validation fails
            form = QuestionnaireForm(request.POST)
            error_message = 'Names must only contain letters and spaces.'
            return render(request, 'survey/questionnaire_form.html', {'form': form, 'error_message': error_message})
        
        # If validation passes, proceed with creating UserResponse
        user_response = UserResponse(
            first_name=first_name,
            last_name=last_name,
            date=request.POST.get('date'),
            staff_name=staff_name,
            sex=request.POST.get('sex'),
            doh_employee=request.POST.get('doh_employee'),
            team_acronym=request.POST.get('team_acronym'),
            bureau_service_region_acronym=request.POST.get('bureau_service_region_acronym'),
            activity_name=request.POST.get('activity_name'),
            expectation=request.POST.get('expectation'),
            statement1_rating=request.POST.get('statement1'),
            statement2_rating=request.POST.get('statement2'),
            statement3_rating=request.POST.get('statement3'),
            statement4_rating=request.POST.get('statement4'),
            statement5_rating=request.POST.get('statement5'),
            statement6_rating=request.POST.get('statement6'),
            quality=request.POST.get('quality'),
            comments=request.POST.get('comments'),
            # Add more fields as needed
        )
        
        # Save the UserResponse instance to the database
        user_response.save()
        
        return redirect('confirmation')  # Redirect to a confirmation page upon successful form submission
    else:
        form = QuestionnaireForm()
    
    statement_fields = [field for field in form if field.name.startswith('statement')]
    
    return render(request, 'survey/questionnaire_form.html', {'form': form, 'statement_fields': statement_fields})

def confirmation(request):
    return render(request, 'survey/confirmation.html')

def landing_page(request):
    # Your logic here, if any
    return render(request, 'survey/landing_page.html')

@login_required(login_url='/admin/login/')
def dashboard(request):
    # return render(request, 'survey/dashboard.html')
    if request.user.is_staff:
        return render(request, 'survey/userresponse_list.html')
    else:
        # Return a '403 Forbidden' response if the user is not an admin
        return HttpResponseForbidden("You are not authorized to view this page.")

from django.shortcuts import get_object_or_404

def qrcode(request):
    return render(request, 'survey/qrcode.html')

def qr(request):
    return render(request, 'survey/qr.html')

def dashboard_view(request):
    # Fetch data from UserResponse model
    responses = UserResponse.objects.all()

    # Count number of responses for each activity
    activity_counts = {}
    for response in responses:
        activity_name = response.activity_name
        if activity_name in activity_counts:
            activity_counts[activity_name] += 1
        else:
            activity_counts[activity_name] = 1

    # Convert activity_counts to lists for Chart.js
    labels = list(activity_counts.keys())
    values = list(activity_counts.values())

    context = {
        'labels': labels,
        'values': values,
        'responses': responses
    }
    return render(request, 'survey/dashboard.html', context)

class Echo:
    """An object that implements just the write method of the file-like
    interface.
    """
    def write(self, value):
        """Write the value by returning it, instead of storing in a buffer."""
        return value

class CSVGenerator:
    """A generator that yields CSV rows."""
    
    def __init__(self, queryset):
        self.queryset = queryset
    
    def __iter__(self):
        # Write header row
        yield ','.join([
            'ID', 'First Name', 'Last Name', 'Date', 'Staff Name',
            'Sex', 'DOH Employee', 'Team Acronym', 'Bureau/Region Acronym',
            'Activity Name', 'Expectation', 'Statement1 Rating', 'Statement2 Rating',
            'Statement3 Rating', 'Statement4 Rating', 'Statement5 Rating',
            'Statement6 Rating', 'Quality', 'Comments'
        ]) + '\n'
        
        # Write data rows
        for response in self.queryset:
            row = [
                str(response.id), response.first_name or '',
                response.last_name or '', str(response.date) if response.date else '',
                response.staff_name or '', response.sex or '',
                response.doh_employee or '', response.team_acronym or '',
                response.bureau_service_region_acronym or '', response.activity_name or '',
                response.expectation or '', str(response.statement1_rating) if response.statement1_rating else '',
                str(response.statement2_rating) if response.statement2_rating else '', str(response.statement3_rating) if response.statement3_rating else '',
                str(response.statement4_rating) if response.statement4_rating else '', str(response.statement5_rating) if response.statement5_rating else '',
                str(response.statement6_rating) if response.statement6_rating else '', response.quality or '', response.comments or ''
            ]
            yield ','.join(row) + '\n'

class UserResponseListView(ListView):
    model = UserResponse
    template_name = 'survey/userresponse_list.html'
    context_object_name = 'responses'
    paginate_by = 10  # Number of records per page

    def get_queryset(self):
        queryset = super().get_queryset().order_by('id')
        return queryset

    def export_to_excel(self, queryset):
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="user_responses.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "User Responses"

        headers = [
            'ID', 'First Name', 'Last Name', 'Date', 'Staff Name',
            'Sex', 'DOH Employee', 'Team Acronym', 'Bureau/Region Acronym',
            'Activity Name', 'Expectation', 'Statement1 Rating',
            'Statement2 Rating', 'Statement3 Rating', 'Statement4 Rating',
            'Statement5 Rating', 'Statement6 Rating', 'Quality', 'Comments'
        ]

        for col_num, header in enumerate(headers, 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            ws[f"{col_letter}1"] = header

        for row_num, obj in enumerate(queryset, 2):
            ws[f"A{row_num}"] = obj.id
            ws[f"B{row_num}"] = obj.first_name
            ws[f"C{row_num}"] = obj.last_name
            ws[f"D{row_num}"] = obj.date
            ws[f"E{row_num}"] = obj.staff_name
            ws[f"F{row_num}"] = obj.sex
            ws[f"G{row_num}"] = obj.doh_employee
            ws[f"H{row_num}"] = obj.team_acronym
            ws[f"I{row_num}"] = obj.bureau_service_region_acronym
            ws[f"J{row_num}"] = obj.activity_name
            ws[f"K{row_num}"] = obj.expectation
            ws[f"L{row_num}"] = obj.statement1_rating
            ws[f"M{row_num}"] = obj.statement2_rating
            ws[f"N{row_num}"] = obj.statement3_rating
            ws[f"O{row_num}"] = obj.statement4_rating
            ws[f"P{row_num}"] = obj.statement5_rating
            ws[f"Q{row_num}"] = obj.statement6_rating
            ws[f"R{row_num}"] = obj.quality
            ws[f"S{row_num}"] = obj.comments

        wb.save(response)
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        paginator = Paginator(self.get_queryset(), self.paginate_by)
        page = self.request.GET.get('page')

        try:
            responses = paginator.page(page)
        except PageNotAnInteger:
            responses = paginator.page(1)
        except EmptyPage:
            responses = paginator.page(paginator.num_pages)
        
        context['responses'] = responses

        try:
            context['is_paginated'] = self.paginate_by < self.get_queryset().count()
        except TypeError:
            context['is_paginated'] = False

        return context

    def get(self, request, *args, **kwargs):
        if 'export' in request.GET:
            queryset = self.get_queryset()
            return self.export_to_excel(queryset)

        return super().get(request, *args, **kwargs)